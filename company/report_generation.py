import os
from wsgiref.util import FileWrapper
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response
from transporter.models import CollectionEntries
from disposal_agency.models import DisposalEntries
from company.models import Company
from .constants import CUSTOMER_REPORTS_COLUMNS
from digitalplatformbackend.constants import user_constants, company_constants
from random import randint
import logging
from datetime import datetime

logger = logging.getLogger(__name__)
from django.shortcuts import redirect
import datetime

def generate_customer_reports(request, **kwarks):
    if kwarks:
        start_date=kwarks['start_date']
        end_date=kwarks['end_date']
    else:
        start_date=None
        end_date=None
    user = request.user
    collection_data_qs = ''
    disposal_data_qs = ''
    file_name = "order" + str(randint(1000000, 9999999)) + "report" + str(randint(1000000, 9999999))
    wb = Workbook()
    ws = wb.active
    headings = [CUSTOMER_REPORTS_COLUMNS.get(col) for col in CUSTOMER_REPORTS_COLUMNS]
    ws.append(headings)
    if start_date and end_date:
        if user.user_type == user_constants.Admin or user.user_type == user_constants.Verifier:
            try:
                collection_data_qs = CollectionEntries.objects.filter(dispatch_date__gte=start_date, dispatch_date__lte=end_date)
                disposal_data_qs = DisposalEntries.objects.filter(unloading_date__gte=start_date, unloading_date__lte=end_date)
            except:
                pass
        elif user.user_type == user_constants.Company_Admin:
            try:
                user_agency=Company.objects.get(id=user.company.id)
                if (user_agency.company_type==company_constants.collection_agency):
                    collection_data_qs = CollectionEntries.objects.filter(company=user.company,dispatch_date__gte=start_date, dispatch_date__lte=end_date)
                disposal_data_qs = DisposalEntries.objects.filter(company=user.company,unloading_date__gte=start_date, unloading_date__lte=end_date)
            except:
                pass
        else:
            try:
                collection_data_qs = CollectionEntries.objects.filter(consignment__order_number__customer_id=user.id,dispatch_date__gte=start_date, dispatch_date__lte=end_date)
                disposal_data_qs = DisposalEntries.objects.filter(consignment__order_number__customer_id=user.id,unloading_date__gte=start_date, unloading_date__lte=end_date)
            except:
                pass
    data = []
    if collection_data_qs:
        disposal_data_qs = DisposalEntries.objects.all()
        for collection_entry in collection_data_qs:
            temp = []
            for col in CUSTOMER_REPORTS_COLUMNS:
                if col == 'order_number':
                    temp.append(
                        collection_entry.consignment.order_number.order_number if collection_entry.consignment.order_number else '-')
                if col == 'consignment_id':
                    temp.append(
                        collection_entry.consignment.consignment_id if collection_entry.consignment.consignment_id else '-')
                if col == 'collection_agency':
                    temp.append(collection_entry.company if collection_entry.company else '-')
                if col == 'vehicle_number':
                    temp.append(collection_entry.vehicle_number if collection_entry.vehicle_number else '-')
                if col == 'Collection_weight':
                    temp.append(str(collection_entry.weight)  if collection_entry.weight else '-')
                if col == 'dispatch_date':
                    temp.append((collection_entry.dispatch_date).strftime("%d-%m-%Y %H:%M:%S") if collection_entry.dispatch_date else '-')
                if col == 'latitude':
                    temp.append(collection_entry.latitude if collection_entry.latitude else '-')
                if col == 'longitude':
                    temp.append(collection_entry.longitude if collection_entry.longitude else '-')
                if col == 'collection_status':
                    if collection_entry.status == 1:
                        temp.append("Approved")
                    elif collection_entry.status == 2:
                        temp.append("Rejected")
                    else:
                        temp.append("Pending")
                if disposal_data_qs:
                    for disposal_entry in disposal_data_qs:
                        if collection_entry.consignment == disposal_entry.consignment:
                            if col == 'disposal_agency': temp.append(disposal_entry.company if disposal_entry.company else '-')
                            if col == 'disposal_weight': temp.append(str(disposal_entry.weight) if disposal_entry.weight else '-')
                            if col == 'unloading_date': temp.append((disposal_entry.unloading_date).strftime("%d-%m-%Y %H:%M:%S") if disposal_entry.unloading_date else '-')
                            if col == 'disposal_latitude': temp.append(disposal_entry.latitude if disposal_entry.latitude else '-')
                            if col == 'disposal_longitude': temp.append(disposal_entry.longitude if disposal_entry.longitude else '-')
                            if col == 'disposal_status':
                                if not disposal_entry.disposal_status:
                                    temp.append('-')
                                elif disposal_entry.disposal_status == 1:
                                    temp.append("Approved")
                                elif disposal_entry.disposal_status == 2:
                                    temp.append("Rejected")
                                else:
                                    temp.append("Pending")
            data.append(temp)
    elif disposal_data_qs:
        collection_data_qs = CollectionEntries.objects.all()
        for collection_entry in collection_data_qs:
            temp = []
            for disposal_entry in disposal_data_qs:
                if disposal_entry.consignment == collection_entry.consignment:
                    for col in CUSTOMER_REPORTS_COLUMNS:
                        if col == 'order_number':
                            temp.append(
                                collection_entry.consignment.order_number if collection_entry.consignment.order_number else '-')
                        if col == 'consignment_id':
                            temp.append(collection_entry.consignment.consignment_id if collection_entry.consignment.consignment_id else '-')
                        if col == 'collection_agency':
                            temp.append(collection_entry.company if collection_entry.company else '-')
                        if col == 'vehicle_number':
                            temp.append(collection_entry.vehicle_number if collection_entry.vehicle_number else '-')
                        if col == 'Collection_weight':
                            temp.append(str(collection_entry.weight) if collection_entry.weight else '-')
                        if col == 'dispatch_date':
                            temp.append((collection_entry.dispatch_date).strftime("%d-%m-%Y %H:%M:%S") if collection_entry.dispatch_date else '-')
                        if col == 'latitude':
                            temp.append(collection_entry.latitude if collection_entry.latitude else '-')
                        if col == 'longitude':
                            temp.append(collection_entry.longitude if collection_entry.longitude else '-')
                        if col == 'collection_status':
                            if collection_entry.status == 1:
                                temp.append("Approved")
                            elif collection_entry.status == 2:
                                temp.append("Rejected")
                            else:
                                temp.append("Pending")
                        if col == 'disposal_agency':
                            temp.append(disposal_entry.company if disposal_entry.company else '-')
                        if col == 'disposal_weight':
                            temp.append(str(disposal_entry.weight) if disposal_entry.weight else '-')
                        if col == 'unloading_date':
                            temp.append((disposal_entry.unloading_date).strftime("%d-%m-%Y %H:%M:%S") if disposal_entry.unloading_date else '-')
                        if col == 'disposal_latitude':
                            temp.append(disposal_entry.latitude if disposal_entry.latitude else '-')
                        if col == 'disposal_longitude':
                            temp.append(disposal_entry.longitude if disposal_entry.longitude else '-')
                        if col == 'disposal_status':
                            if not disposal_entry.disposal_status:
                                temp.append("-")
                            elif disposal_entry.disposal_status == 1:
                                temp.append("Approved")
                            elif disposal_entry.disposal_status == 2:
                                temp.append("Rejected")
                            else:
                                temp.append("Pending")
                    data.append(temp)
    for row in data:
        str_arr = []
        for col in row:
            str_arr.append(str(col))
        ws.append(str_arr)
    # SET WIDTH OF ALL COLUMNS AS PER DATA
    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if cell.value and len(cell.value) > column_widths[i]:
                    column_widths[i] = len(cell.value) + 2
            else:
                if cell.value:
                    column_widths += [len(cell.value) + 2]
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width
    file_name = file_name
    xlsx_path = 'static/reports/{0}.xlsx'.format(file_name)
    font_format = Font(size=12, bold=True)
    for cell in ws["1:1"]:
        cell.font = font_format
    wb.save(xlsx_path)
    file_path = xlsx_path
    return file_wrapper(file_path)


def file_wrapper(file_path):

    wrapper = FileWrapper(open(file_path, 'rb'))
    response = HttpResponse(wrapper, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={0}'.format(os.path.basename(file_path))
    os.remove(file_path)
    logger.warning(f"report Generated successfully")
    return response
